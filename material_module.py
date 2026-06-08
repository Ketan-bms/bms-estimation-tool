"""
material_module.py
──────────────────
Material estimate module for BMS Estimation Tool.
Embeds into app.py as module_material(p) call inside the Estimate tab.

Price book structure:
  PRICEBOOKS dict: {name → pricebook_json}
  Stored in st.session_state.pricebooks

Per-project material estimate stored in p["material"]:
  {"items": [{description, manufacturer, part_no, qty, unit_cost, ext_cost, section, subsection}],
   "pricebook": "Honeywell Standard"}
"""

import json, io
from copy import deepcopy
from collections import defaultdict
import streamlit as st
import pandas as pd

# ── Embedded Honeywell price book (200 items) ─────────────────────────────────
def _honeywell_pricebook():
    import os
    pb_path = os.path.join(os.path.dirname(__file__) if "__file__" in dir() else ".", "pricebook_honeywell.json")
    try:
        with open(pb_path) as f:
            return json.load(f)
    except Exception:
        pass
    # Minimal fallback if file not found
    return {
        "name": "Honeywell Standard",
        "manufacturer": "Honeywell",
        "sections": [
            {"section": "DDC Controllers", "subsections": [
                {"subsection": "BACnet/IP Controllers - CPO", "items": [
                    {"description": "CPO Plant Controller", "manufacturer": "Honeywell", "part_no": "CPO-PC400", "protocol": "BACNET IP", "unit_cost": 1610.00},
                    {"description": "CPO BACnet Router Controller", "manufacturer": "Honeywell", "part_no": "CP-CORE", "protocol": "BACNET IP", "unit_cost": 953.04},
                    {"description": "Cipher 50 Plant Controller (No Display)", "manufacturer": "Honeywell", "part_no": "WEB-EHSERIESNX26ND", "protocol": "BACNET IP", "unit_cost": 761.62},
                    {"description": "JACE Controller 250 pts", "manufacturer": "Honeywell", "part_no": "WEB-8005", "protocol": "BACNET IP", "unit_cost": 1026.23},
                    {"description": "JACE Controller 500 pts", "manufacturer": "Honeywell", "part_no": "WEB-8010", "protocol": "BACNET IP", "unit_cost": 1185.17},
                    {"description": "CPO Room Controller (10UI 6AO 4DO)", "manufacturer": "Honeywell", "part_no": "CPO-RL5", "protocol": "BACNET MS/TP", "unit_cost": 315.18},
                ]}
            ]},
            {"section": "Field Instrumentation", "subsections": [
                {"subsection": "Wall Mount Sensors", "items": [
                    {"description": "Wall Module, Temp Only", "manufacturer": "Honeywell", "part_no": "TR21/U", "unit_cost": 15.73},
                    {"description": "Wall Module, Temp/Humidity", "manufacturer": "Honeywell", "part_no": "TR21-H/U", "unit_cost": 108.80},
                    {"description": "CO2 Sensor, Duct Mount", "manufacturer": "Honeywell", "part_no": "C7232B1006", "unit_cost": 412.00},
                    {"description": "Current Switch - Split Core", "manufacturer": "Senva", "part_no": "C-2345", "unit_cost": 25.00},
                ]}
            ]},
            {"section": "Actuators", "subsections": [
                {"subsection": "Belimo Modulating", "items": [
                    {"description": "35 IN-LB Damper Actuator", "manufacturer": "Belimo", "part_no": "LF24-SR", "unit_cost": 191.10},
                    {"description": "90 IN-LB Damper Actuator", "manufacturer": "Belimo", "part_no": "NFB24-SR", "unit_cost": 209.70},
                    {"description": "180 IN-LB Damper Actuator", "manufacturer": "Belimo", "part_no": "AFB24-SR", "unit_cost": 242.40},
                ]}
            ]},
            {"section": "Panel & Hardware", "subsections": [
                {"subsection": "Control Panels - Standard", "items": [
                    {"description": "Control Panel Normal (36x30) - Complete", "manufacturer": "Hoffman/Symcon", "part_no": "A363008LP kit", "unit_cost": 1583.00},
                    {"description": "Control Panel Small (30x24) - Complete", "manufacturer": "Hoffman/Symcon", "part_no": "A302408LP kit", "unit_cost": 1818.00},
                ]}
            ]},
        ]
    }


def init_pricebooks():
    if "pricebooks" not in st.session_state:
        st.session_state.pricebooks = {
            "Honeywell Standard": _honeywell_pricebook()
        }


def all_items_flat(pb):
    """Return flat list of all items with section/subsection labels."""
    rows = []
    for sec in pb.get("sections", []):
        for sub in sec.get("subsections", []):
            for item in sub.get("items", []):
                rows.append({
                    "section":    sec["section"],
                    "subsection": sub["subsection"],
                    **item,
                    "qty": 0,
                    "ext_cost": 0.0,
                })
    return rows


# ── Main material estimate UI ─────────────────────────────────────────────────
def module_material(p):
    init_pricebooks()

    if "material" not in p:
        p["material"] = {"items": [], "pricebook": "Honeywell Standard", "markup": 10}

    mat = p["material"]
    pricebooks = st.session_state.pricebooks

    st.markdown("### Material estimate")
    st.caption("Select items from the price book, enter quantities. "
               "AI can suggest materials based on your takeoff and point list.")

    # ── Top controls ──────────────────────────────────────────────────────────
    hc1, hc2, hc3, hc4 = st.columns([2, 1, 1, 1])

    pb_names = list(pricebooks.keys())
    selected_pb = hc1.selectbox("Price book", pb_names,
                                index=pb_names.index(mat["pricebook"])
                                      if mat["pricebook"] in pb_names else 0,
                                key="mat_pb_sel")
    mat["pricebook"] = selected_pb
    pb = pricebooks[selected_pb]

    mat["markup"] = hc2.number_input("Markup %", 0, 100,
                                      int(mat.get("markup", 10)), key="mat_markup")

    if hc3.button("🤖 AI suggest materials", key="ai_mat"):
        api_key = st.session_state.get("anthropic_api_key",
                                        __import__("os").environ.get("ANTHROPIC_API_KEY",""))
        if not api_key:
            st.error("Add API key in sidebar.")
        else:
            with st.spinner("Claude is selecting materials from the price book..."):
                suggestions = _ai_suggest_materials(p, pb, api_key)
            if suggestions:
                # Merge suggestions into existing items
                existing_parts = {i["part_no"]: i for i in mat["items"]}
                for s in suggestions:
                    pn = s.get("part_no","")
                    if pn in existing_parts:
                        existing_parts[pn]["qty"] = max(
                            existing_parts[pn].get("qty",0), s.get("qty",0))
                    else:
                        mat["items"].append(s)
                st.success(f"✅ {len(suggestions)} items suggested by AI.")
                st.rerun()

    if hc4.button("Clear all", key="mat_clear"):
        mat["items"] = []
        st.rerun()

    st.divider()

    # ── Two-column layout: price book browser | selected items ────────────────
    col_browser, col_selected = st.columns([1.2, 1])

    with col_browser:
        st.markdown("**Price book — browse & add items**")

        # Search + section filter
        sr1, sr2 = st.columns([1.5, 1])
        search = sr1.text_input("Search", placeholder="JACE, actuator, CO2, enclosure…",
                                 key="mat_search")
        sections = ["All sections"] + [s["section"] for s in pb.get("sections", [])]
        sec_filter = sr2.selectbox("Section", sections, key="mat_sec_filter")

        # Build filtered item list
        all_items = all_items_flat(pb)
        filtered = all_items
        if search:
            sl = search.lower()
            filtered = [i for i in filtered
                        if sl in i["description"].lower()
                        or sl in i.get("part_no","").lower()
                        or sl in i.get("manufacturer","").lower()]
        if sec_filter != "All sections":
            filtered = [i for i in filtered if i["section"] == sec_filter]

        st.caption(f"{len(filtered)} items")

        # Group by subsection for display
        by_sub = defaultdict(list)
        for item in filtered:
            by_sub[f"{item['section']} › {item['subsection']}"].append(item)

        for group_name, items in list(by_sub.items())[:20]:  # limit display
            with st.expander(group_name, expanded=len(by_sub) <= 3):
                for item in items:
                    ic1, ic2, ic3, ic4 = st.columns([3, 1.2, 0.8, 0.8])
                    ic1.markdown(f"**{item['description']}**")
                    ic1.caption(f"`{item.get('part_no','')}` · {item.get('manufacturer','')}")
                    ic2.markdown(f"${item['unit_cost']:,.2f}")
                    qty = ic3.number_input("Qty", 0, 9999, 0,
                                           key=f"mat_qty_{item['part_no']}_{item['description'][:20]}")
                    if ic4.button("Add →", key=f"mat_add_{item['part_no']}_{item['description'][:15]}"):
                        if qty > 0:
                            _add_item(mat, item, qty)
                            st.rerun()

        if len(by_sub) > 20:
            st.caption(f"Showing first 20 groups. Refine your search to see more.")

    with col_selected:
        st.markdown("**Selected items**")

        items = mat.get("items", [])
        if not items:
            st.info("No items selected yet. Browse the price book on the left and add items.")
        else:
            subtotal = 0.0
            remove_idx = None

            for i, item in enumerate(items):
                ext = item.get("qty", 0) * item.get("unit_cost", 0)
                item["ext_cost"] = ext
                subtotal += ext

                with st.container(border=True):
                    r1, r2 = st.columns([3, 1])
                    r1.markdown(f"**{item['description']}**")
                    r1.caption(f"`{item.get('part_no','')}` · {item.get('manufacturer','')}")
                    r2.markdown(f"${ext:,.2f}")

                    q1, q2, q3 = st.columns([1, 1, 1])
                    new_qty = q1.number_input("Qty", 0, 9999,
                                               int(item.get("qty", 0)),
                                               key=f"sel_qty_{i}_{item.get('part_no','')}",
                                               label_visibility="collapsed")
                    item["qty"] = new_qty
                    item["ext_cost"] = new_qty * item.get("unit_cost", 0)
                    q2.caption(f"@ ${item['unit_cost']:,.2f}/ea")
                    if q3.button("✕", key=f"rm_{i}", help="Remove"):
                        remove_idx = i

            if remove_idx is not None:
                mat["items"].pop(remove_idx)
                st.rerun()

            # Recalculate subtotal
            subtotal = sum(i.get("qty",0) * i.get("unit_cost",0) for i in items)
            markup_pct = mat.get("markup", 10)
            markup_amt = subtotal * markup_pct / 100
            grand_total = subtotal + markup_amt

            st.divider()
            t1, t2, t3 = st.columns(3)
            t1.metric("Material subtotal", f"${subtotal:,.0f}")
            t2.metric(f"Markup ({markup_pct}%)", f"${markup_amt:,.0f}")
            t3.metric("Material total", f"${grand_total:,.0f}")

            st.divider()
            if st.button("⬇ Export material estimate to Excel", key="exp_mat"):
                xb = _export_material_excel(items, subtotal, markup_pct, markup_amt,
                                             grand_total, p["name"], selected_pb)
                st.download_button(
                    "Download .xlsx", xb,
                    f"material_{p['name'].replace(' ','_')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_mat"
                )

    st.divider()
    # ── Add custom item ───────────────────────────────────────────────────────
    with st.expander("➕ Add custom item (not in price book)"):
        with st.form("custom_item_form"):
            ci1, ci2, ci3 = st.columns([3, 1.5, 1])
            c_desc = ci1.text_input("Description *")
            c_mfr  = ci1.text_input("Manufacturer")
            c_part = ci2.text_input("Part number")
            c_cost = ci2.number_input("Unit cost $", 0.0, 999999.0, 0.0, step=0.01)
            c_qty  = ci3.number_input("Qty", 0, 9999, 1)
            c_sec  = ci3.selectbox("Section",
                                    [s["section"] for s in pb.get("sections",[])] + ["Other"])
            if st.form_submit_button("Add custom item"):
                if c_desc:
                    mat["items"].append({
                        "description": c_desc,
                        "manufacturer": c_mfr,
                        "part_no": c_part,
                        "unit_cost": float(c_cost),
                        "qty": int(c_qty),
                        "ext_cost": float(c_cost) * int(c_qty),
                        "section": c_sec,
                        "subsection": "Custom",
                    })
                    st.success("Added."); st.rerun()

    st.divider()
    # ── Price book management ─────────────────────────────────────────────────
    with st.expander("⚙️ Manage price books"):
        st.markdown("**Add a new price book** (upload your own JSON or Excel)")
        pb1, pb2 = st.columns([1, 1])
        new_pb_name = pb1.text_input("Price book name", placeholder="Johnson Controls 2024")
        new_pb_file = pb2.file_uploader("Upload JSON price book", type=["json"], key="pb_upload")
        if st.button("Save price book", key="save_pb"):
            if new_pb_name and new_pb_file:
                try:
                    pb_data = json.load(new_pb_file)
                    pb_data["name"] = new_pb_name
                    st.session_state.pricebooks[new_pb_name] = pb_data
                    st.success(f"✅ '{new_pb_name}' saved.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not parse JSON: {e}")

        st.divider()
        st.markdown("**Download current price book as JSON** (to edit and re-upload)")
        pb_json = json.dumps(pb, indent=2)
        st.download_button("⬇ Download price book JSON",
                           pb_json.encode(),
                           f"pricebook_{selected_pb.replace(' ','_')}.json",
                           "application/json", key="dl_pb")

        if len(pricebooks) > 1:
            del_pb = st.selectbox("Delete price book",
                                   [n for n in pricebooks if n != "Honeywell Standard"],
                                   key="del_pb_sel")
            if st.button("Delete", key="del_pb_btn"):
                del st.session_state.pricebooks[del_pb]
                st.rerun()


def _add_item(mat, item, qty):
    # Check if already in list — update qty
    for existing in mat["items"]:
        if existing.get("part_no") == item.get("part_no") and \
           existing.get("description") == item.get("description"):
            existing["qty"] = existing.get("qty", 0) + qty
            existing["ext_cost"] = existing["qty"] * existing["unit_cost"]
            return
    mat["items"].append({
        "description":  item["description"],
        "manufacturer": item.get("manufacturer", ""),
        "vendor":       item.get("vendor", ""),
        "part_no":      item.get("part_no", ""),
        "protocol":     item.get("protocol", ""),
        "section":      item.get("section", ""),
        "subsection":   item.get("subsection", ""),
        "unit_cost":    item["unit_cost"],
        "qty":          qty,
        "ext_cost":     qty * item["unit_cost"],
    })


def _ai_suggest_materials(p, pb, api_key):
    """AI reads takeoff + point list and suggests material quantities from the price book."""
    equip  = p["takeoff"].get("equipment", [])
    points = p["point_list"].get("rows", [])

    # Build concise catalog for AI
    catalog = []
    for sec in pb.get("sections", []):
        for sub in sec.get("subsections", []):
            for item in sub.get("items", []):
                catalog.append({
                    "part_no":     item.get("part_no",""),
                    "description": item["description"],
                    "unit_cost":   item["unit_cost"],
                    "section":     sec["section"],
                    "subsection":  sub["subsection"],
                    "manufacturer":item.get("manufacturer",""),
                    "vendor":      item.get("vendor",""),
                })

    # Count point types from point list
    pt_counts = {"AI":0,"AO":0,"DI":0,"DO":0,"HWI":0,"Network":0}
    for row in points:
        for k in pt_counts:
            if str(row.get(k,"")).strip() == "1":
                pt_counts[k] += 1

    # Device summary
    dev_summary = {}
    for e in equip:
        sys = e.get("system","Unknown")
        dev_summary[sys] = dev_summary.get(sys,0) + 1

    prompt = (
        f"You are a BMS estimator selecting materials for project '{p['name']}'.\n\n"
        f"Device summary: {json.dumps(dev_summary)}\n"
        f"Point counts: {json.dumps(pt_counts)}\n"
        f"Total devices: {len(equip)}, Total points: {len(points)}\n\n"
        f"From the following price book catalog, select the most appropriate items and "
        f"estimate realistic quantities for this project.\n"
        f"Price book (first 80 items shown): {json.dumps(catalog[:80])}\n\n"
        f"Return ONLY a JSON array of selected items:\n"
        f'[{{"part_no":"CPO-RL5","description":"CPO Room Controller SYLK BUS","manufacturer":"Honeywell",'
        f'"vendor":"Honeywell","unit_cost":315.18,"qty":8,"section":"DDC Controllers",'
        f'"subsection":"Unitary Controllers - CPO","ext_cost":2521.44}}]\n\n'
        f"Select 10-20 items. Focus on: controllers (qty based on device count), "
        f"sensors (qty based on point count), actuators (qty based on damper/valve points), "
        f"one appropriately-sized control panel, wiring accessories.\n"
        f"Return ONLY the JSON array."
    )

    try:
        import anthropic
        msg = anthropic.Anthropic(api_key=api_key).messages.create(
            model="claude-sonnet-4-6", max_tokens=2000,
            messages=[{"role":"user","content":prompt}]
        )
        raw = msg.content[0].text.strip()
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"): raw = raw[4:]
        return json.loads(raw.strip())
    except Exception as e:
        st.error(f"AI suggestion error: {e}")
        return []


def _export_material_excel(items, subtotal, markup_pct, markup_amt, grand_total,
                            proj_name, pb_name):
    """Export material estimate to Excel matching the format of the original price book."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Material Estimate"

    # Header
    ws.merge_cells("A1:J1")
    ws["A1"] = f"BMS Material Estimate — {proj_name}"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Price book: {pb_name}"
    ws["A2"].font = Font(italic=True, size=10, name="Arial")

    # Column headers (row 4)
    headers = ["Section","Subsection","Description","Manufacturer","Part No.",
               "Protocol","Qty","Unit Cost","Ext Cost","Notes"]
    col_widths = [18,22,40,18,18,16,6,12,12,20]
    hdr_fill = PatternFill("solid", start_color="1F4E79")

    for ci, (h,w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(4, ci, h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[4].height = 24

    # Group items by section
    by_section = defaultdict(list)
    for item in items:
        by_section[item.get("section","Other")].append(item)

    row = 5
    sec_fill   = PatternFill("solid", start_color="D6E4F0")
    alt_fill   = PatternFill("solid", start_color="F8FBFE")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    section_subtotals = {}

    for sec_name, sec_items in by_section.items():
        # Section header row
        ws.merge_cells(f"A{row}:J{row}")
        ws.cell(row, 1, sec_name).font = Font(bold=True, name="Arial", size=11)
        ws.cell(row, 1).fill = sec_fill
        ws.row_dimensions[row].height = 18
        row += 1

        sec_total = 0
        for i, item in enumerate(sec_items):
            fill = alt_fill if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            ext = item.get("qty",0) * item.get("unit_cost",0)
            sec_total += ext
            row_data = [
                item.get("section",""),
                item.get("subsection",""),
                item.get("description",""),
                item.get("manufacturer",""),
                item.get("part_no",""),
                item.get("protocol",""),
                item.get("qty",0),
                item.get("unit_cost",0),
                ext,
                item.get("notes",""),
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row, ci, val)
                cell.fill = fill
                cell.border = border
                cell.font = Font(name="Arial", size=10)
                if ci in (7,8,9):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            row += 1

        # Section subtotal
        ws.cell(row, 8, f"{sec_name} Subtotal").font = Font(bold=True, name="Arial", size=10)
        ws.cell(row, 8).alignment = Alignment(horizontal="right")
        ws.cell(row, 9, sec_total).font = Font(bold=True, name="Arial", size=10)
        ws.cell(row, 9).number_format = '$#,##0.00'
        ws.cell(row, 9).fill = PatternFill("solid", start_color="E8F4F8")
        section_subtotals[sec_name] = sec_total
        row += 2

    # Totals block
    tot_fill = PatternFill("solid", start_color="1F4E79")
    for label, val in [
        ("Material Subtotal", subtotal),
        (f"Markup ({markup_pct}%)", markup_amt),
        ("MATERIAL TOTAL", grand_total),
    ]:
        ws.cell(row, 8, label).font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        ws.cell(row, 8).fill = tot_fill
        ws.cell(row, 8).alignment = Alignment(horizontal="right")
        ws.cell(row, 9, val).font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        ws.cell(row, 9).fill = tot_fill
        ws.cell(row, 9).number_format = '$#,##0.00'
        ws.cell(row, 9).alignment = Alignment(horizontal="right")
        row += 1

    ws.freeze_panes = "A5"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
