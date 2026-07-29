"""
soo_module_ui.py
Complete SOO extraction UI module with 5 buttons:
1. Overview
2. Proposal
3. Point List (main)
4. Appendix
5. Important Notes

Drop this into app.py or call module_soo(project_dict) from main
"""

import streamlit as st
import pandas as pd
import json
from io import BytesIO


def _extract_text_from_pdf(pdf_bytes, max_chars=15000):
    """Extract text from PDF using PyMuPDF."""
    try:
        import fitz
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        text = "".join(page.get_text() for page in doc)
        return text[:max_chars]
    except Exception as e:
        st.error(f"PDF extraction failed: {e}")
        return ""


def _extract_text_from_docx(docx_bytes, max_chars=15000):
    """Extract text from DOCX."""
    try:
        from docx import Document
        doc = Document(BytesIO(docx_bytes))
        text = "\n".join(p.text for p in doc.paragraphs)
        return text[:max_chars]
    except Exception as e:
        st.error(f"DOCX extraction failed: {e}")
        return ""


def _claude(api_key, prompt, max_tokens=3000):
    """Call Claude API."""
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-opus-4.5",
            max_tokens=max_tokens,
            messages=[{"role": "user", "content": prompt}]
        )
        return response.content[0].text
    except Exception as e:
        st.error(f"Claude API error: {e}")
        return None


def api_key():
    """Get API key from Streamlit secrets."""
    return st.secrets.get("ANTHROPIC_API_KEY", "")


def module_soo(p):
    """
    SOO (Sequence of Operations) extraction module.
    5 extraction buttons: Overview, Proposal, Point List, Appendix, Important Notes
    """
    
    st.markdown("### 📋 Sequence of Operations (SOO) Processing")
    st.markdown("""
    Upload your SOO PDF/DOCX and extract:
    - **Overview:** Bird's eye view of control approach per system
    - **Proposal:** Generate DOCX proposal (you provide template)
    - **Point List:** Main BMS points in Excel format (11 columns)
    - **Appendix:** Special sequences (fire safety, emergency, future)
    - **Important Notes:** Key estimation points
    """)
    
    # ── File Upload ────────────────────────────────────────────────────────────
    uploaded_file = st.file_uploader("Upload SOO PDF or DOCX", type=["pdf", "docx"], key="soo_upload")
    
    if not uploaded_file:
        st.info("👆 Upload a SOO document to begin analysis")
        return
    
    # ── Extract SOO Text ───────────────────────────────────────────────────────
    soo_bytes = uploaded_file.read()
    fname = uploaded_file.name.lower()
    
    if fname.endswith(".docx"):
        soo_text = _extract_text_from_docx(soo_bytes, 15000)
    else:
        soo_text = _extract_text_from_pdf(soo_bytes, 15000)
    
    if not soo_text:
        st.error("Could not extract text from SOO document")
        return
    
    st.success(f"✅ Loaded: {uploaded_file.name} ({len(soo_text)} chars)")
    
    # Initialize SOO state
    if "soo_data" not in p:
        p["soo_data"] = {
            "overview": None,
            "proposal": None,
            "point_list": None,
            "appendix": None,
            "important_notes": None
        }
    
    # ── 5 Extraction Buttons ───────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### Extract from SOO:")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    k = api_key()
    if not k:
        st.error("❌ No API key. Set ANTHROPIC_API_KEY in Streamlit Secrets.")
        return
    
    # ── Button 1: Overview ─────────────────────────────────────────────────────
    with col1:
        if st.button("📋 Overview", key="btn_overview", use_container_width=True):
            with st.spinner("Extracting overview..."):
                from soo_extractor import generate_overview_prompt
                prompt = generate_overview_prompt(p.get("name", "Project"), soo_text)
                raw = _claude(k, prompt, max_tokens=2000)
                if raw:
                    try:
                        data = json.loads(raw[raw.find("{"):raw.rfind("}")+1])
                        p["soo_data"]["overview"] = data
                        st.success("✅ Overview extracted")
                    except:
                        st.warning("Could not parse response")
    
    # ── Button 2: Proposal ─────────────────────────────────────────────────────
    with col2:
        if st.button("📄 Proposal", key="btn_proposal", use_container_width=True):
            st.info("🔗 Upload your proposal template (DOCX) in the sidebar to generate proposal")
            # This would require template uploader - shown separately below
    
    # ── Button 3: Point List ───────────────────────────────────────────────────
    with col3:
        if st.button("📊 Point List", key="btn_pointlist", use_container_width=True):
            with st.spinner("Generating point list..."):
                from soo_extractor import generate_pointlist_prompt, parse_pointlist_response
                prompt = generate_pointlist_prompt(p.get("name", "Project"), soo_text)
                raw = _claude(k, prompt, max_tokens=4000)
                if raw:
                    try:
                        rows = parse_pointlist_response(raw)
                        p["soo_data"]["point_list"] = rows
                        st.success(f"✅ {len(rows)} points extracted")
                    except Exception as e:
                        st.error(f"Parse error: {e}")
    
    # ── Button 4: Appendix ─────────────────────────────────────────────────────
    with col4:
        if st.button("📎 Appendix", key="btn_appendix", use_container_width=True):
            if not p["soo_data"]["point_list"]:
                st.warning("⚠️ Generate Point List first")
            else:
                with st.spinner("Generating appendix..."):
                    from soo_extractor import generate_appendix_prompt, parse_pointlist_response
                    main_equip = list(set(row.get("Equipment", "") for row in p["soo_data"]["point_list"]))
                    prompt = generate_appendix_prompt(p.get("name", "Project"), soo_text, main_equip)
                    raw = _claude(k, prompt, max_tokens=2000)
                    if raw:
                        try:
                            rows = parse_pointlist_response(raw)
                            p["soo_data"]["appendix"] = rows
                            st.success(f"✅ {len(rows)} appendix points extracted")
                        except Exception as e:
                            st.error(f"Parse error: {e}")
    
    # ── Button 5: Important Notes ──────────────────────────────────────────────
    with col5:
        if st.button("⭐ Notes", key="btn_notes", use_container_width=True):
            with st.spinner("Extracting notes..."):
                from soo_extractor import generate_important_notes_prompt, parse_notes_response
                prompt = generate_important_notes_prompt(p.get("name", "Project"), soo_text)
                raw = _claude(k, prompt, max_tokens=2500)
                if raw:
                    try:
                        data = parse_notes_response(raw)
                        p["soo_data"]["important_notes"] = data
                        st.success("✅ Important notes extracted")
                    except Exception as e:
                        st.error(f"Parse error: {e}")
    
    # ── Display Results ────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### Results:")
    
    # Overview
    if p["soo_data"]["overview"]:
        with st.expander("📋 Overview (Bird's eye view)"):
            overview_data = p["soo_data"]["overview"]
            if isinstance(overview_data, dict) and "overview" in overview_data:
                for system in overview_data["overview"]:
                    st.write(f"**{system.get('System', 'Unknown')}** — {system.get('Equipment_Type', '')}")
                    st.write(f"Control: {system.get('Control_Approach', '')}")
                    st.write(f"Points: {system.get('Control_Points', '')}")
                    st.write(f"Integration: {system.get('Integration', '')}")
                    st.write("---")
            else:
                st.json(overview_data)
    
    # Point List
    if p["soo_data"]["point_list"]:
        with st.expander(f"📊 Point List ({len(p['soo_data']['point_list'])} points)"):
            df = pd.DataFrame(p["soo_data"]["point_list"])
            # Reorder columns if needed
            cols = ["Panel Name", "Equipment", "Point name", "Control Device", 
                   "AI", "BI", "AO", "BO", "Serial Pt", "Terms", "Remarks"]
            for col in cols:
                if col not in df.columns:
                    df[col] = ""
            df = df[cols]
            
            st.dataframe(df, use_container_width=True, height=400)
            
            # Export to Excel
            xb = BytesIO()
            df.to_excel(xb, index=False, sheet_name="Point List")
            xb.seek(0)
            st.download_button("📥 Download Point List Excel", xb, 
                              f"point_list_{p.get('name', 'project').replace(' ', '_')}.xlsx",
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    # Appendix
    if p["soo_data"]["appendix"]:
        with st.expander(f"📎 Appendix ({len(p['soo_data']['appendix'])} special points)"):
            df = pd.DataFrame(p["soo_data"]["appendix"])
            cols = ["Panel Name", "Equipment", "Point name", "Control Device", 
                   "AI", "BI", "AO", "BO", "Serial Pt", "Terms", "Remarks"]
            for col in cols:
                if col not in df.columns:
                    df[col] = ""
            df = df[cols]
            
            st.dataframe(df, use_container_width=True, height=300)
            
            # Export to Excel
            xb = BytesIO()
            df.to_excel(xb, index=False, sheet_name="Appendix")
            xb.seek(0)
            st.download_button("📥 Download Appendix Excel", xb,
                              f"appendix_{p.get('name', 'project').replace(' ', '_')}.xlsx",
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    # Combined export (Main + Appendix)
    if p["soo_data"]["point_list"] and p["soo_data"]["appendix"]:
        st.markdown("**Or export both together:**")
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        
        # Main sheet
        ws_main = wb.active
        ws_main.title = "Main Points"
        
        df_main = pd.DataFrame(p["soo_data"]["point_list"])
        cols = ["Panel Name", "Equipment", "Point name", "Control Device", 
               "AI", "BI", "AO", "BO", "Serial Pt", "Terms", "Remarks"]
        for col in cols:
            if col not in df_main.columns:
                df_main[col] = ""
        df_main = df_main[cols]
        
        fill = PatternFill("solid", start_color="2E75B6")
        for ci, col in enumerate(cols, 1):
            cell = ws_main.cell(1, ci, col)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = fill
        
        for ri, row in enumerate(df_main.to_dict("records"), 2):
            for ci, col in enumerate(cols, 1):
                ws_main.cell(ri, ci, row.get(col, ""))
        
        # Appendix sheet
        ws_app = wb.create_sheet("Appendix")
        
        df_app = pd.DataFrame(p["soo_data"]["appendix"])
        for col in cols:
            if col not in df_app.columns:
                df_app[col] = ""
        df_app = df_app[cols]
        
        fill_app = PatternFill("solid", start_color="B85C00")
        for ci, col in enumerate(cols, 1):
            cell = ws_app.cell(1, ci, col)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = fill_app
        
        for ri, row in enumerate(df_app.to_dict("records"), 2):
            for ci, col in enumerate(cols, 1):
                ws_app.cell(ri, ci, row.get(col, ""))
        
        xb_combined = BytesIO()
        wb.save(xb_combined)
        xb_combined.seek(0)
        
        st.download_button("📥 Download Main + Appendix Combined", xb_combined,
                          f"point_list_complete_{p.get('name', 'project').replace(' ', '_')}.xlsx",
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    # Important Notes
    if p["soo_data"]["important_notes"]:
        with st.expander("⭐ Important Notes (Estimation)"):
            notes = p["soo_data"]["important_notes"]
            if isinstance(notes, dict):
                for category, items in notes.items():
                    if items:  # Only show non-empty categories
                        st.subheader(category.replace("_", " ").title())
                        for item in items:
                            st.write(f"• {item}")
            else:
                st.json(notes)
    
    # ── Proposal (Separate Section) ────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### Generate Proposal from SOO")
    
    proposal_template = st.file_uploader("📄 Your proposal template (DOCX)", 
                                        type=["docx"], key="proposal_template")
    
    if proposal_template and st.button("Generate Proposal", key="btn_gen_proposal"):
        with st.spinner("Generating proposal..."):
            from proposal_generator_module import generate_proposal_prompt
            template_bytes = proposal_template.read()
            
            prompt = generate_proposal_prompt(
                p.get("name", "Project"),
                soo_text,
                None,  # template_analysis (can be enhanced)
                p.get("client", "")
            )
            raw = _claude(k, prompt, max_tokens=3500)
            if raw:
                # Generate DOCX from text
                # For now, just offer text download (DOCX generation requires python-docx setup)
                st.text_area("Generated Proposal (copy to Word):", raw, height=400)
                st.download_button("📥 Download as Text", raw, 
                                  f"proposal_{p.get('name', 'project')}.txt")


if __name__ == "__main__":
    # Test
    test_project = {"name": "Test", "client": ""}
    module_soo(test_project)
